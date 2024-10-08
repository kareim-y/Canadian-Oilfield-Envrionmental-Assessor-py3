#python_to_OPGEE
from model_inputs import ModelInputs
import pickle

def colnum_string(n):
    #covnverts number n to the column string - ie C = 3, AB = 28
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def python_to_OPGEE(OPGEE_data):

	with open('model_input_instance.pkl', 'rb') as f:
		inputs_instance = pickle.load(f)

	import openpyxl
	from string import ascii_uppercase
	import re
	from .map_to_drive import map_to_drive #path to Project Data folder

	#import_file_name = "OPGEE_v2.0_Original.xlsm"
	#import_file_name = ""OPGEE_3.0a_BETA.xlsm""

	# import_export_files = ["OPGEE_v2.0_tight_oil_edit.xlsm"]
	import_export_files = ["OPGEE_3.0c_BETA.xlsm"]
	#,"OPGEE_3.0a_BETA_edit.xlsm"

	for file in import_export_files:

		# Kareem edits: 
		# old code
		# path = map_to_drive() + "Project Data/OPGEE/" + file
		# new code
		path = "Project Data/OPGEE/" + file
		
		#these can occur if a well is drilled in 2016 and produces in 2017
		#if we are only assessing the period to the end of 2016
		excluded_wells = [] 
		exported_wells = [] 

		print('\n\nWriting data to Excel OPGEE.......\n')
		print(('Original File Location:  \n\n' + str(path)))
		print('\n')

		print(' ------------------- Exclude Exported Wells ------------------- \n')
		print('Exclude wells to reduce/remove calculation issues in OPGEE')
		print('Oil production must be greater than 0 bbl/day for OPGEE to run\n')

		# Kareem Edits
		min_age = float(inputs_instance.min_welltime)
		print('Chosen Response for \'Minimum well producing time (years)\' is:', inputs_instance.min_welltime)
		min_oil_prod = float(inputs_instance.min_wellprod)
		print('Chosen Response for \'Minimum oil production (bbl/day)\' is:', inputs_instance.min_wellprod)
		print('\n')

		wb=openpyxl.load_workbook(filename = path, read_only=False, keep_vba=True)
		sheetlist = wb.sheetnames

		#-----------OPGEE_input Sheet data ----------------

		#finding the right sheet  
		for sheet in sheetlist:
			if sheet == 'Inputs':
				inputs_sheet = wb[sheet]

		if __name__ == "__main__": #we only want to do this on test run throughs
			field_assessed = 'defaults'
		else:
			field_assessed = 'assessed field'
		
		#getting the appropriate start point

		initial_column_number = 8
		column_number = initial_column_number # (A = 0) 8 equates to I using colnum_string - this is where we start out input data

		for entry in OPGEE_data:
			#if entry in ['assessed field']: #just export the assessed field
			if entry not in ['headings','defaults','units','excel position']:
				#want to make sure the entered wells are oil producing 
				age = OPGEE_data[entry][OPGEE_data['headings'].index('Field age')]
				oil_prod = OPGEE_data[entry][OPGEE_data['headings'].index('Oil production volume')]
				if (oil_prod >= min_oil_prod) and (age >= min_age):
					exported_wells.append(entry)
					for i in range(0,len(OPGEE_data[entry])):
						#we want to ensure the wells have produced oil, otherwis OPGEE will have errors runnning
						column = colnum_string(column_number)
						row = str(OPGEE_data['excel position'][i])
						position = column + row
						inputs_sheet[position] = OPGEE_data[entry][i]
					column_number = column_number + 1
				
				else:
					excluded_wells.append(entry)

		# Specify the rows and starting column
		row_87 = 87
		row_86 = 86
		row_46 = 46
		start_column = 'H'

		# Initialize column index
		column_index = openpyxl.utils.column_index_from_string(start_column)
		# Loop through the cells in row 87 and perform the equation
		while True:
			cell_87 = inputs_sheet.cell(row=row_87, column=column_index)
			if cell_87.value is None:  # Stop when you reach a blank cell
				break

			cell_46 = inputs_sheet.cell(row=row_46, column=column_index)
			cell_86 = inputs_sheet.cell(row=row_86, column=column_index)

			# Perform the calculation
			original_value = cell_87.value
			if cell_46.value == 0 :
				new_value = 0
			else:
				new_value = original_value / (cell_46.value - cell_86.value)

			# Store the original value in the row below
			inputs_sheet.cell(row=row_87 + 2, column=column_index, value=original_value)

			# Replace the value in row 87 with the new value
			cell_87.value = new_value

			# Move to the next column
			column_index += 1

		#we use the project name we specify at the start

		field_name_index = OPGEE_data['headings'].index('Field name')
		project_name = OPGEE_data['assessed field'][field_name_index]
		remove_characters = ['/', ' ']
		project_name = re.sub("|".join(remove_characters), "", project_name) 
		
		opgee_version = file[0:11]
		#opgee_version = "OPGEE_3.0a_BETA_"
		#opgee_version = file[:-5] + '_'

		export_file_name =  opgee_version + project_name + ".xlsm"

		# Kareem edits:
		# old code
		# file_save_location = map_to_drive() +"Project Data/OPGEE/COEA - OPGEE/" + export_file_name
		# new code
		file_save_location = "Project Data/OPGEE/COEA - OPGEE/" + export_file_name

		wb.save(file_save_location)

		print('Successfully Exported!\n')
		print((str(len(exported_wells)) + ' Have been exported to OPGEE'))
		print(('\nFile Location:  \n' + str(file_save_location)))
		
		print(('\n' + str(len(excluded_wells)) + ' Wells have been excluded'))
		print('\nExclusions')
		print(excluded_wells)
		print('\n')
		print(('# Wells producing less than ' + str(min_oil_prod) + ' bbl/day'))
		print(('# Less than ' + str(min_age) + ' producing years'))

		wb.close()

	return

if __name__ == "__main__":

	from .OPGEE_defaults import OPGEE_defaults
	from .OPGEE_drilling_and_development import OPGEE_drilling_and_development
	from .well_search import well_search
	from .general_well_data_analysis import OPGEE_well_data, general_well_data_analysis
	
	OPGEE_data = OPGEE_defaults()

	#well_data_headings, well_data = get_formation_well_data()# MONTNEY
	well_data_headings, well_data, project_name = well_search()

	project_name = 'Trial'

	OPGEE_data = OPGEE_defaults()

	OPGEE_data = general_well_data_analysis(well_data_headings, well_data, OPGEE_data, project_name)

	#OPGEE_data = OPGEE_well_data(well_data, well_data_headings, OPGEE_data)

	#OPGEE_data = OPGEE_drilling_and_development(OPGEE_data, well_data, well_data_headings)
	
	python_to_OPGEE(OPGEE_data)
